"""
Phase 1 pipeline adapter — the only module that imports ml_package.

The worker thread captures stdout, so print() banners in run_phase1 flow
straight into the live log box. STAGE_PROGRESS keys match those banners
verbatim — the worker's _LogStream lowercases each line and substring-
matches to advance the progress bar.
"""

from __future__ import annotations

import os
import threading
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from typing import Any, Optional

import openpyxl
import pandas as pd

from ml_package import ensemble as _ens
from ml_package import mapping_lookup as _ml
from ml_package import text_match as _tm
from ml_package import ml_classifier as _rfx
from ml_package.write_results import write_results


STAGE_PROGRESS = {
    "read inputs":   0.10,
    "mappinglookup": 0.22,
    "textmatch":     0.68,
    "ensemble":      0.78,
    "qc_ready":      0.85,
    "write output":  0.98,
    "done":          1.00,
}

STAGE_LABELS = {
    "read inputs":   "Reading & validating input files…",
    "mappinglookup": "Running exact & fuzzy lookup matching…",
    "textmatch":     "Predicting attributes for unresolved products (lookup + ML)…",
    "ensemble":      "Combining lookup + ML results and assigning QC priority…",
    "qc_ready":      "Pipeline complete — QC review ready",
    "write output":  "Writing output Excel workbook…",
    "done":          "Done",
}

# Pre-flight coverage check (Lookup-only) progress + labels. Same banner keys
# as the full run so the worker's _LogStream can drive the bar.
PRECHECK_PROGRESS = {"read inputs": 0.20, "mappinglookup": 0.80}
PRECHECK_LABELS = {
    "read inputs":   "Reading inputs…",
    "mappinglookup": "Checking lookup coverage…",
}


class PipelineStopped(Exception):
    """Raised between stages when the user has requested a stop."""


@dataclass
class Phase1Payload:
    FINAL: pd.DataFrame
    FLAT_FILE_OUT: Any
    meta: pd.DataFrame
    dictEnsemble: dict[str, pd.DataFrame]


def run_phase1(excel_path: str, csv_path: str,
               stop_event: Optional[threading.Event] = None) -> Phase1Payload:
    """
    Phase 1 pipeline: read inputs → mappinglookup → BM25 + XGBoost in
    parallel → ensemble. Prints stage banners that the worker's _LogStream
    parses to advance the progress bar. Honours stop_event between stages.
    """
    def _maybe_stop():
        if stop_event and stop_event.is_set():
            raise PipelineStopped()

    # ── Read inputs ───────────────────────────────────────────────────────
    print("START read inputs")
    print(f"  Reading Excel: {os.path.basename(excel_path)}")
    wb          = openpyxl.load_workbook(excel_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    print(f"  Excel sheets found: {sheet_names}")

    meta_sheet  = next((s for s in sheet_names if "META"  in s.upper()), None)
    final_sheet = next((s for s in sheet_names if "FINAL" in s.upper()), None)
    if not meta_sheet:
        raise RuntimeError(f"No META sheet in {excel_path}. Found: {sheet_names}")
    if not final_sheet:
        raise RuntimeError(f"No FINAL sheet in {excel_path}. Found: {sheet_names}")

    print(f"  Loading META sheet ({meta_sheet}) and FINAL sheet ({final_sheet})…")
    metaGridPDF    = pd.read_excel(excel_path, sheet_name=meta_sheet)
    combinedAMPPDF = pd.read_excel(excel_path, sheet_name=final_sheet)
    print(f"  Reading CSV: {os.path.basename(csv_path)}")
    attrGridPDF    = pd.read_csv(csv_path, low_memory=False)
    print(f"  Loaded — META {metaGridPDF.shape} | FINAL {combinedAMPPDF.shape} | CSV {attrGridPDF.shape}")

    # Strip '|' (pipe) from DESCRIPTION — some client submission systems treat it
    # as a delimiter and reject the file. Replace with a space so adjacent words
    # don't merge.
    for _df in (combinedAMPPDF, attrGridPDF):
        if "DESCRIPTION" in _df.columns:
            _df["DESCRIPTION"] = _df["DESCRIPTION"].astype(str).str.replace("|", " ", regex=False).str.strip()

    FINAL   = combinedAMPPDF.copy()
    SAL_COL = "RAW_TOTAL_DOLLARS"
    if SAL_COL not in combinedAMPPDF.columns:
        SAL_COL = next((c for c in combinedAMPPDF.columns if "dollar" in c.lower()), None)
        print(f"  WARNING: using {SAL_COL!r} as sales column")
    print("DONE read inputs")
    _maybe_stop()

    # ── Validate + prep META ──────────────────────────────────────────────
    for col in ("Attribute Name in MDM", "Attribute Group name", "Attribute_Type", "Type"):
        if col not in metaGridPDF.columns:
            raise RuntimeError(f"META sheet missing column: '{col}'")

    metaGridPDF_old = metaGridPDF.copy()
    # Case-insensitive match: META sheets export Attribute_Type as "Modeling"/
    # "Reporting" (title case), not "MODELING". Strip + upper before comparing.
    metaGridPDF     = metaGridPDF[
        metaGridPDF["Attribute_Type"].astype(str).str.strip().str.upper() == "MODELING"
    ]
    metaFields      = list(set(
        list(metaGridPDF["Attribute Name in MDM"].unique()) +
        list(metaGridPDF["Attribute Group name"].unique())
    ))

    missing_flat = list(set(metaGridPDF["Attribute Name in MDM"]) - set(attrGridPDF.columns))
    missing_amp  = list(set(metaFields) - set(combinedAMPPDF.columns))
    if missing_flat:
        print(f"  WARNING - columns missing from CSV: {missing_flat}")
    if missing_amp:
        print(f"  WARNING - columns missing from FINAL: {missing_amp}")

    attrGridPDF    = attrGridPDF.astype(object).fillna("nan").astype(str)
    combinedAMPPDF = combinedAMPPDF.astype(object).fillna("nan").astype(str)
    metaGridPDF    = metaGridPDF.astype(object).fillna("nan").astype(str)
    combinedAMPPDF["TOTAL_UNIT_SALES"] = pd.to_numeric(
        combinedAMPPDF[SAL_COL], errors="coerce"
    ).fillna(0)

    dictRecom: dict = {}

    # ── Lookup ────────────────────────────────────────────────────────────
    print("START mappinglookup")
    print("  Exact & fuzzy lookup — finds direct matches for new products using values already seen in the historical labelled data.")
    n_attrs = len(metaGridPDF["Attribute Name in MDM"].unique())
    n_new   = len(attrGridPDF)
    print(f"  Matching {n_new} new products against {n_attrs} attributes…")
    dictRecom, _, FLAT_FILE_OUT, dict_split_parent = _ml.runLookup(
        attrGridPDF, metaGridPDF, combinedAMPPDF, dictRecom
    )
    print("DONE mappinglookup")
    _maybe_stop()

    # ── BM25 + XGBoost in parallel ────────────────────────────────────────
    print("START textmatch")
    print("  For products not resolved by lookup, predicting attributes from patterns in your historical data (lookup + ML)…")
    _dictRecom_tm: dict = {}
    _dictRecom_ml: list = [None]

    def _run_tm():
        if stop_event and stop_event.is_set():
            return
        _tm.runTextMatch(metaGridPDF, combinedAMPPDF, attrGridPDF, _dictRecom_tm)

    def _run_ml():
        if stop_event and stop_event.is_set():
            return
        _dictRecom_ml[0] = _rfx.runML(attrGridPDF, metaGridPDF, combinedAMPPDF)

    with ThreadPoolExecutor(max_workers=4) as pool:
        f_tm = pool.submit(_run_tm)
        f_ml = pool.submit(_run_ml)
        f_tm.result()
        f_ml.result()

    dictRecom.update(_dictRecom_tm)
    if _dictRecom_ml[0]:
        dictRecom.update(_dictRecom_ml[0])
    print("DONE textmatch")
    _maybe_stop()

    # ── Ensemble ──────────────────────────────────────────────────────────
    print("START ensemble")
    print("  Combining ML predictions with lookup results and assigning QC priority…")
    dictEnsemble = _ens.runEnsemble(dictRecom, metaGridPDF, dict_split_parent)
    print("DONE ensemble")
    _maybe_stop()

    print(f"DONE pipeline — {len(dictEnsemble)} lookup sheet(s) ready for QC review: {', '.join(dictEnsemble.keys())}")
    return Phase1Payload(
        FINAL=FINAL,
        FLAT_FILE_OUT=FLAT_FILE_OUT,
        meta=metaGridPDF_old,
        dictEnsemble=dictEnsemble,
    )


def _raise_if_stopped(stop_event: Optional[threading.Event]) -> None:
    if stop_event is not None and stop_event.is_set():
        raise PipelineStopped()


def _extract_coverage_gaps(recom_dict: dict, sample_cap: int = 25) -> list[dict]:
    """
    Derive coverage gaps from the Lookup tables: rows written blank with
    score 0 are flat-file key combos that found no historical match (the
    exact condition _build_lookup_table reports). Returns one entry per
    affected attribute, busiest first.
    """
    gaps: list[dict] = []
    for key, lkp in recom_dict.items():
        if not key.startswith("Lookup_") or not hasattr(lkp, "columns"):
            continue
        attr = key[len("Lookup_"):]
        if attr not in lkp.columns or "score" not in lkp.columns:
            continue
        key_cols = [c for c in lkp.columns if c not in (attr, "score", "Rank", "Record")]
        if not key_cols:
            continue
        total_combos = int(lkp[key_cols].drop_duplicates().shape[0])
        value = lkp[attr].astype(str).str.strip()
        score = pd.to_numeric(lkp["score"], errors="coerce")
        gap_rows = lkp.loc[(value == "") & (score == 0), key_cols].drop_duplicates()
        if gap_rows.empty:
            continue
        combos = [
            ", ".join(f"{c}={str(row[c])!r}" for c in key_cols)
            for _, row in gap_rows.iterrows()
        ]
        gaps.append({
            "attribute":     attr,
            "key_columns":   list(key_cols),
            "count":         int(len(gap_rows)),     # unmatched (unseen-value) combos
            "total_combos":  total_combos,           # total distinct flat-file combos for the attr
            "sample_combos": combos[:sample_cap],
        })
    gaps.sort(key=lambda g: g["count"], reverse=True)
    return gaps


def run_coverage_check(excel_path: str, csv_path: str,
                       stop_event: Optional[threading.Event] = None) -> list[dict]:
    """
    Pre-flight gate: run ONLY the Lookup stage and report attributes whose
    flat-file key combos have no historical match — the cheap (~1 min) signal
    that lets the UI warn the analyst and offer continue/cancel BEFORE the full
    multi-minute run. Mirrors run_phase1's input prep so the gate sees exactly
    what the real run will. Returns [{attribute, key_columns, count, sample_combos}].
    """
    print("START read inputs")
    print(f"  Reading Excel: {os.path.basename(excel_path)}")
    wb          = openpyxl.load_workbook(excel_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    meta_sheet  = next((s for s in sheet_names if "META"  in s.upper()), None)
    final_sheet = next((s for s in sheet_names if "FINAL" in s.upper()), None)
    if not meta_sheet or not final_sheet:
        raise RuntimeError(
            f"Workbook needs META and FINAL sheets. Found: {sheet_names}"
        )
    metaGridPDF    = pd.read_excel(excel_path, sheet_name=meta_sheet)
    combinedAMPPDF = pd.read_excel(excel_path, sheet_name=final_sheet)
    attrGridPDF    = pd.read_csv(csv_path, low_memory=False)
    print("DONE read inputs")
    _raise_if_stopped(stop_event)

    # Mirror run_phase1's prep: filter META to MODELING, string-coerce all
    # frames, and derive TOTAL_UNIT_SALES so runLookup ranks by sales volume.
    SAL_COL = "RAW_TOTAL_DOLLARS"
    if SAL_COL not in combinedAMPPDF.columns:
        SAL_COL = next((c for c in combinedAMPPDF.columns if "dollar" in c.lower()), None)
    metaGridPDF = metaGridPDF[
        metaGridPDF["Attribute_Type"].astype(str).str.strip().str.upper() == "MODELING"
    ]
    attrGridPDF    = attrGridPDF.astype(object).fillna("nan").astype(str)
    combinedAMPPDF = combinedAMPPDF.astype(object).fillna("nan").astype(str)
    metaGridPDF    = metaGridPDF.astype(object).fillna("nan").astype(str)
    combinedAMPPDF["TOTAL_UNIT_SALES"] = (
        pd.to_numeric(combinedAMPPDF[SAL_COL], errors="coerce").fillna(0)
        if SAL_COL else 0
    )

    print("START mappinglookup")
    print("  Pre-flight coverage check — Lookup only; finding key combos with no historical match.")
    recom_dict: dict = {}
    _ml.runLookup(attrGridPDF, metaGridPDF, combinedAMPPDF, recom_dict)
    print("DONE mappinglookup")
    _raise_if_stopped(stop_event)

    gaps = _extract_coverage_gaps(recom_dict)
    if gaps:
        total = sum(g["count"] for g in gaps)
        print(f"Pre-flight: {total} key combo(s) across {len(gaps)} attribute(s) "
              f"have no historical match — see warnings above.")
    else:
        print("Pre-flight: no coverage gaps — every flat-file key combo matched history.")
    return gaps


def write_qc_excel(out_path: str, payload: Phase1Payload,
                   qc_edits: dict[str, pd.DataFrame]) -> None:
    """
    Write File_For_Mapping_QC.xlsx using analyst-edited lookup DataFrames.

    Sheets present in `qc_edits` override the originals from the pipeline;
    any key not in qc_edits falls back to payload.dictEnsemble[key].
    """
    final_dict = {
        key: qc_edits.get(key, payload.dictEnsemble[key])
        for key in payload.dictEnsemble.keys()
    }
    write_results(out_path, payload.FINAL, payload.FLAT_FILE_OUT, payload.meta, final_dict)
